from datetime import datetime
from io import BytesIO

import numpy as np
import openpyxl
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import quote_sheetname, column_index_from_string
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.dimensions import ColumnDimension

from utils import common


def excelExport(title, data):
    # 创建工作簿和工作表
    workbook = Workbook()
    worksheet = workbook.active

    format_data = data_format(title, data)
    set_export_title(worksheet, format_data['title_label'])
    set_export_value(worksheet, format_data['newData'])

    # 冻结第一行
    worksheet.freeze_panes = 'A2'  # 从A列开始冻结第一行
    # 设置行颜色

    # 保存工作簿
    # workbook.save("output.xlsx")
    excel_file = BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)
    workbook.close()
    return excel_file

def excelExportModel(title, choices):
    # 创建工作簿和工作表
    workbook = Workbook()
    worksheet = workbook.active

    format_data = data_format(title, [])
    set_export_title(worksheet, format_data['title_label'])
    if len(choices) > 0:
        set_export_choices(worksheet, choices)

    # 冻结第一行
    worksheet.freeze_panes = 'A2'  # 从A列开始冻结第一行
    # 设置行颜色

    # 保存工作簿
    excel_file = BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)
    workbook.close()
    return excel_file


def set_export_title(worksheet, title):
    # 设置样式
    font = Font(color="000000", bold=True, size=15)
    fill = PatternFill(fill_type="solid", fgColor="8DB4E2")
    # 创建Border对象
    border = Border(
        left=Side(border_style="thin", color="000000"),  # 左边框
        right=Side(border_style="thin", color="000000"),  # 右边框
        top=Side(border_style="thin", color="000000"),  # 上边框
        bottom=Side(border_style="thin", color="000000")  # 下边框
    )
    # 循环设置列的宽度、样式和内容
    for index, item in enumerate(title):
        column_letter = openpyxl.utils.get_column_letter(index + 1)
        worksheet.column_dimensions[column_letter].width = 5 * len(item)
        cell = worksheet.cell(row=1, column=index + 1)
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.value = item


def set_export_value(worksheet, data):
    font = Font(color="000000", size=12)
    fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    fill2 = PatternFill(fill_type="solid", fgColor="FFFFFF")
    # 创建Border对象
    border = Border(
        left=Side(border_style="thin", color="000000"),  # 左边框
        right=Side(border_style="thin", color="000000"),  # 右边框
        top=Side(border_style="thin", color="000000"),  # 上边框
        bottom=Side(border_style="thin", color="000000")  # 下边框
    )
    index = 1
    for row in data:
        index += 1

        for row_index, item in enumerate(row):
            # 获取行对象并设置颜色
            cell = worksheet.cell(row=index, column=row_index + 1)
            if index % 2 == 0:
                cell.fill = fill
            else:
                cell.fill = fill2
            cell.border = border
            cell.font = font
            cell.value = row[item]


def set_export_choices(worksheet, choices):
    # 创建DataValidation对象
    # 遍历整个A列并添加数据验证规则
    for item in choices:
        if item.get("array", False):
            dv = DataValidation(type="list", formula1=f'"{",".join(item["array"])}"')
            column_letter = item['column']
            column_index = column_index_from_string(column_letter)
            max = item.get("max", 5000)
            for row in worksheet.iter_rows(min_row=2, max_row=max, min_col=column_index, max_col=column_index):
                for cell in row:
                    dv.add(cell)

            # 在工作表中添加数据验证
            worksheet.add_data_validation(dv)
        if item.get("format", False):
            column_letter = item['column']
            column_index = column_index_from_string(column_letter)
            max = item['max']
            for row in worksheet.iter_rows(min_row=2, max_row=max, min_col=column_index, max_col=column_index):
                for cell in row:
                    cell.number_format = item["format"]


def convert_to_excel_column(number):
    column = ""
    while number > 0:
        number -= 1
        column = chr((number % 26) + 65) + column
        number //= 26
    return column


def get_choices_dv(choices_array, row_index):
    dv = DataValidation(type="list", formula1='"{}"'.format(",".join(choices_array[row_index])),
                        showDropDown=True)
    return dv


def data_format(title, data):
    title_label = []
    newData = []
    for item in title:
        title_label.append(item['value'])
    for item in data:
        tem_data = {}
        for titem in title:
            if titem.get("type", False):
                if titem['type'] == 'date':
                    tem_data[titem['key']] = common.getTimeStr(item[titem['key']])
                    continue
            tem_data[titem['key']] = item[titem['key']]
        newData.append(tem_data)

    return {'title_label': title_label, 'newData': newData}


def import_excel(file, header_mapping, model):
    excel_data = pd.read_excel(file)
    for index, row in excel_data.iterrows():
        obj = model()
        for header, django_field in header_mapping.items():
            if row.get(header, False):
                if type(row[header]) == str:
                    setattr(obj, django_field, row[header])

                elif not np.isnan(float(row[header])):
                    setattr(obj, django_field, row[header])
        obj.save()


def import_excel2(request, header_mapping):
    if request.method == 'POST' and request.FILES['file']:
        file = request.FILES['file']
        df = pd.read_excel(file)
        excel_array = df.values.tolist()
        num_keys = len(header_mapping)
        rel = []
        for row in excel_array:
            new_list = {header_mapping[i]['key']: row[i] for i in range(num_keys)}
            rel.append(new_list)
        return rel
def import_excel_falsk(file, header_mapping):
    df = pd.read_excel(file)
    excel_array = df.values.tolist()
    num_keys = len(header_mapping)
    rel = []
    for row in excel_array:
        new_list = {header_mapping[i]['key']: row[i] for i in range(num_keys)}
        rel.append(new_list)
    return rel



